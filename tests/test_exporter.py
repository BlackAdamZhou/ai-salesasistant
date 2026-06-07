from io import BytesIO

from openpyxl import load_workbook

from app.exporter import build_analysis_workbook


REQUIRED_SHEETS = {
    "Summary_结论",
    "门店表现",
    "商品销售额",
    "动销速度",
    "备货建议",
    "低动销商品",
    "日期分析",
    "区域表现",
}


def _summary():
    return {
        "row_count": 2,
        "product_count": 1,
        "store_count": 1,
        "has_stock_column": True,
        "date_range": {
            "start_date": "2026-04-01",
            "end_date": "2026-04-02",
            "sales_days": 2,
        },
        "store_performance": [
            {
                "store_name": "Store A",
                "region": "Region 1",
                "total_sales_amount": 100,
                "total_quantity_sold": 10,
                "transaction_count": 2,
                "sales_days": 2,
                "average_daily_sales": 50,
                "best_sales_date": "2026-04-01",
                "best_day_sales": 60,
            }
        ],
        "region_performance": [
            {
                "region": "Region 1",
                "total_sales_amount": 100,
                "total_quantity_sold": 10,
                "store_count": 1,
                "transaction_count": 2,
                "sales_days": 2,
                "average_daily_sales": 50,
            }
        ],
        "top_products": [
            {
                "product_code": "Product_001",
                "total_sales_amount": 100,
                "total_quantity_sold": 10,
                "sales_frequency": 2,
                "sales_days": 2,
                "covered_store_count": 1,
                "average_sales_amount": 50,
                "average_daily_quantity": 5,
                "sales_velocity": 5,
                "single_store_daily_quantity": 5,
                "average_stock_remaining": 20,
            }
        ],
        "fast_moving_products": [
            {
                "product_code": "Product_001",
                "sales_velocity": 5,
                "total_quantity_sold": 10,
                "total_sales_amount": 100,
                "sales_frequency": 2,
                "sales_days": 2,
                "covered_store_count": 1,
                "single_store_daily_quantity": 5,
            }
        ],
        "slow_moving_products": [],
        "stocking_tiers": {
            "a_plus_core_products": [],
            "b_class_products": [],
            "low_moving_products": [],
        },
        "stocking_recommendations": [
            {
                "product_code": "Product_001",
                "recommendation": "Maintain current stock level",
                "reason": "Stable sales.",
            }
        ],
        "date_sales_relationship": {
            "peak_date": "2026-04-01",
            "peak_date_sales_amount": 60,
            "trend": "stable",
            "weekend_sales_lift_percent": None,
            "correlation_metrics": {"sales_amount": 0.1},
            "top_sales_dates": [{"date": "2026-04-01", "sales_amount": 60}],
            "lowest_sales_dates": [{"date": "2026-04-02", "sales_amount": 40}],
            "weekend_vs_weekday": [{"day_type": "weekday", "day_count": 2}],
            "daily_sales": [
                {
                    "date": "2026-04-01",
                    "sales_amount": 60,
                    "quantity_sold": 6,
                    "transaction_count": 1,
                    "active_store_count": 1,
                    "active_product_count": 1,
                }
            ],
        },
    }


def test_build_analysis_workbook_contains_required_sheets():
    workbook_bytes = build_analysis_workbook(_summary(), ai_report="test report")
    workbook = load_workbook(BytesIO(workbook_bytes.getvalue()))

    assert REQUIRED_SHEETS <= set(workbook.sheetnames)
