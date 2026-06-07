from pathlib import Path

from openpyxl import load_workbook

from app.ai_service import generate_fallback_report
from app.analysis import build_analysis_summary
from app.anonymizer import anonymise_product_names
from app.file_parser import clean_sales_dataframe, read_sales_file
from app.main import _build_ai_safe_summary, _build_analysis_workbook


def test_full_pipeline_with_sample_csv():
    path = Path("data/sample_sales.csv")
    cleaned = read_sales_file(path.name, path.read_bytes())
    anonymised, mapping = anonymise_product_names(cleaned)
    summary = build_analysis_summary(anonymised)
    report = generate_fallback_report(summary)

    assert mapping
    assert summary["store_performance"]
    assert summary["top_products"]
    assert "Which Stores Perform Well" in report
    assert "Coca Cola" not in report
    assert "Jasmine Rice" not in report
    assert "Potato Chips" not in report


def test_ai_safe_summary_is_compact():
    summary = {
        "row_count": 100,
        "product_count": 20,
        "store_count": 30,
        "has_stock_column": False,
        "date_range": {"sales_days": 19},
        "store_performance": [{"store_name": str(index)} for index in range(30)],
        "top_products": [
            {"product_code": f"Product_{index:03d}", "total_sales_amount": index}
            for index in range(30)
        ],
        "fast_moving_products": [{"product_code": str(index)} for index in range(30)],
        "slow_moving_products": [{"product_code": str(index)} for index in range(30)],
        "stocking_recommendations": [
            {"product_code": str(index), "recommendation": "Maintain", "reason": "x"}
            for index in range(30)
        ],
        "stocking_classification": {
            "a_plus_core_products": [{"product_code": str(index)} for index in range(12)],
            "b_class_products": [{"product_code": str(index)} for index in range(12)],
            "low_moving_products": [{"product_code": str(index)} for index in range(12)],
        },
        "date_sales_relationship": {
            "peak_date": "2026-04-01",
            "daily_sales": [{"date": str(index)} for index in range(30)],
            "top_sales_dates": [{"date": "2026-04-01"}],
            "lowest_sales_dates": [{"date": "2026-04-02"}],
        },
    }

    ai_summary = _build_ai_safe_summary(summary)

    assert len(ai_summary["store_performance_top_10_by_sales_amount"]) == 10
    assert len(ai_summary["products_top_10_by_sales_amount"]) == 10
    assert len(ai_summary["stocking_recommendations_sample"]) == 20
    assert ai_summary["has_stock_column"] is False
    assert len(ai_summary["stocking_classification"]["a_plus_core_products"]) == 10
    assert "daily_sales" not in ai_summary["date_sales_relationship"]


def test_clean_sales_dataframe_does_not_convert_missing_text_to_nan_string():
    cleaned = clean_sales_dataframe(
        read_sales_file(
            "sample.csv",
            (
                "日期,门店名称,商品名称,销售数量,销售金额\n"
                "2026-04-01,,匿名商品A,1,10\n"
                "2026-04-01,门店A,匿名商品A,2,20\n"
            ).encode("utf-8"),
        )
    )

    assert len(cleaned) == 1
    assert "nan" not in set(cleaned["store_name"].astype(str).str.lower())


def test_full_pipeline_with_chinese_excel_fixture():
    path = Path("tests/fixtures/sample_pos_zh.xlsx")
    cleaned = read_sales_file(path.name, path.read_bytes())
    anonymised, mapping = anonymise_product_names(cleaned)
    summary = build_analysis_summary(anonymised)
    report = generate_fallback_report(summary, output_language="zh")

    assert mapping
    assert summary["has_stock_column"] is True
    assert summary["store_performance"][0]["total_sales_amount"] > 0
    assert summary["top_products"][0]["total_sales_amount"] >= summary["top_products"][-1]["total_sales_amount"]
    assert "| 排名 | 门店 | 区域 | 销售额 | 销量 | 日均销售额 |" in report
    assert "| 排名 | 商品代码 | 销售额 | 销量 | 覆盖门店 |" in report
    assert "匿名商品甲" not in report
    assert "匿名商品乙" not in report


def test_export_analysis_workbook_contains_expected_sheets():
    path = Path("tests/fixtures/sample_pos_zh.xlsx")
    cleaned = read_sales_file(path.name, path.read_bytes())
    anonymised, _ = anonymise_product_names(cleaned)
    summary = build_analysis_summary(anonymised)
    workbook_stream = _build_analysis_workbook(summary)
    workbook = load_workbook(workbook_stream, read_only=True)

    assert workbook.sheetnames == [
        "Summary_结论",
        "门店表现",
        "商品销售额",
        "动销速度",
        "备货建议",
        "低动销商品",
        "日期分析",
        "区域表现",
    ]
