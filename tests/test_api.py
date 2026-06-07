from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.exporter import XLSX_MEDIA_TYPE
from app.main import app


client = TestClient(app)
REQUIRED_SHEETS = {
    "Summary_结论",
    "门店表现",
    "商品销售额",
    "动销速度",
    "备货建议",
    "低动销商品",
    "日期分析",
    "区域表现",
}


def test_analyze_sales_still_returns_json_with_anonymised_report():
    path = Path("data/sample_sales.csv")

    response = client.post(
        "/analyze-sales",
        data={"ai_provider": "local", "output_language": "zh"},
        files={"file": (path.name, path.read_bytes(), "text/csv")},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["store_count"] > 0
    assert "has_stock_column" in payload
    assert payload["region_performance"]
    assert "Product_" in payload["ai_report"]
    assert "Coca Cola" not in payload["ai_report"]


def test_export_analysis_returns_downloadable_xlsx():
    path = Path("data/sample_sales.csv")

    response = client.post(
        "/export-analysis",
        data={"ai_provider": "local", "output_language": "zh"},
        files={"file": (path.name, path.read_bytes(), "text/csv")},
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == XLSX_MEDIA_TYPE
    assert "filename*=UTF-8''ai_sales_analysis_" in response.headers[
        "content-disposition"
    ]

    workbook = load_workbook(BytesIO(response.content))
    assert REQUIRED_SHEETS <= set(workbook.sheetnames)

    product_sheet_text = "\n".join(
        str(cell.value)
        for row in workbook["商品销售额"].iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "Product_" in product_sheet_text
    assert "Coca Cola" not in product_sheet_text
